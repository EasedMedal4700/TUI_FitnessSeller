import os
import csv
import json
import re
import shutil
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import datetime


def safe_name(name: str) -> str:
    if name is None:
        return "unknown"
    s = re.sub(r"[^A-Za-z0-9_-]", "_", str(name))
    return s or "unknown"


def process_excel(config_path='config.yaml'):
    import yaml
    from openpyxl import load_workbook

    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = yaml.safe_load(f)
    excel_file = cfg['excel_file']
    wb = load_workbook(excel_file)
    try:
        sheet = wb['containers']
    except Exception:
        sheet = wb.active

    rows = []
    for row in sheet.iter_rows(values_only=True):
        converted = []
        for cell in row:
            if isinstance(cell, datetime.datetime):
                converted.append(cell.isoformat())
            else:
                converted.append(cell)
        if any(c is not None for c in converted):
            rows.append(converted)

    out_csv = Path('data') / 'containers_data.csv'
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, 'w', newline='', encoding='utf-8') as cf:
        writer = csv.writer(cf)
        writer.writerows(rows)
    wb.close()
    print(f'Wrote {len(rows)} rows to {out_csv}')
    return out_csv


def parse_tracking(html_text):
    soup = BeautifulSoup(html_text, 'html.parser')
    result = {'tracking': [], 'pod': []}
    # simplified parse: reuse html_screen heuristics
    # find tables with headers mentioning Event and Date
    for table in soup.find_all('table'):
        ths = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        if any('event' in h for h in ths) and any('date' in h or 'time' in h for h in ths):
            headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
            for tr in table.find_all('tr'):
                if tr.find_all('th'):
                    continue
                cols = [td.get_text(strip=True) for td in tr.find_all(['td','th'])]
                if not any(cols):
                    continue
                # map columns heuristically
                rec = {}
                # try positional mapping
                keys = ['event','date_time','operation','location','details']
                for i,k in enumerate(keys):
                    if i < len(cols):
                        rec[k] = cols[i]
                result['tracking'].append(rec)

    # POD: look for tables with 'POD' or 'Received'
    for table in soup.find_all('table'):
        ths = [th.get_text(strip=True).lower() for th in table.find_all('th')]
        if any('received' in h or 'pod' in h or 'comments' in h for h in ths):
            headers = [th.get_text(strip=True).lower() for th in table.find_all('th')]
            for tr in table.find_all('tr'):
                if tr.find_all('th'):
                    continue
                cols = [td.get_text(strip=True) for td in tr.find_all(['td','th'])]
                if not any(cols):
                    continue
                podrec = {'date_time': cols[0] if len(cols)>0 else None,
                          'received_by': cols[1] if len(cols)>1 else None,
                          'comments': cols[2] if len(cols)>2 else None}
                result['pod'].append(podrec)

    return result


def fetch_pages_and_attachments(csv_path):
    html_dir = Path('data') / 'container_HTML'
    html_dir.mkdir(parents=True, exist_ok=True)
    count = 0
    failures = 0
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    for r in rows:
        if not r:
            continue
        container = r[0]
        link = r[-1]
        if not link:
            continue
        link = str(link).strip()
        if not link:
            continue
        filename = safe_name(container)
        html_path = html_dir / f"{filename}.html"
        json_path = html_dir / f"{filename}.json"
        try:
            resp = requests.get(link, timeout=20)
            resp.raise_for_status()
            text = resp.text
            html_path.write_text(text, encoding='utf-8')
            parsed = parse_tracking(text)
            # attachments + anchor logging for debugging
            attachments = []
            soup = BeautifulSoup(text, 'html.parser')
            anchors_log = []
            for a in soup.find_all('a'):
                anchor = a.get_text(separator=' ', strip=True) or ''
                href = a.get('href')
                anchors_log.append({'text': anchor, 'href': href})
                if not href:
                    continue
                if any(k in anchor.lower() for k in ('packing','invoice','bill of lading','download')):
                    file_url = urljoin(link, href)
                    try:
                        rfile = requests.get(file_url, timeout=30, stream=True)
                        rfile.raise_for_status()
                        cd = rfile.headers.get('content-disposition','')
                        if 'filename=' in cd:
                            fname = cd.split('filename=')[-1].strip('"')
                        else:
                            fname = os.path.basename(file_url.split('?',1)[0]) or safe_name(anchor)
                        root, ext = os.path.splitext(fname)
                        if not ext:
                            ct = rfile.headers.get('content-type','')
                            ext = '.pdf' if 'pdf' in ct else '.bin'
                        save_name = f"{filename}_{safe_name(root)}{ext}"
                        save_path = html_dir / save_name
                        with open(save_path, 'wb') as of:
                            for chunk in rfile.iter_content(8192):
                                if chunk:
                                    of.write(chunk)
                        attachments.append(str(save_path))
                    except Exception as e:
                        print(f'attachment download failed for {container}: {e}')
                        continue

            # write anchors log for debugging
            try:
                anchors_path = html_dir / f"{filename}_anchors.json"
                with open(anchors_path, 'w', encoding='utf-8') as af:
                    json.dump(anchors_log, af, ensure_ascii=False, indent=2)
            except Exception:
                pass

            parsed['container'] = container
            parsed['url'] = link
            parsed['attachments'] = attachments
            with open(json_path, 'w', encoding='utf-8') as jf:
                json.dump(parsed, jf, ensure_ascii=False, indent=2)
            count += 1
            print(f'Processed {container}')
        except Exception as e:
            failures += 1
            print(f'Failed {container}: {e}')

    print(f'Pages processed: {count}, failures: {failures}')


if __name__ == '__main__':
    csvp = process_excel()
    fetch_pages_and_attachments(csvp)
    # run ETAs
    from code.extract_eta import extract_etas
    extract_etas()
